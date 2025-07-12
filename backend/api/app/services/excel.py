from pathlib import Path
from typing import Any, Dict, List, Optional
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.cell import column_index_from_string
from openpyxl.worksheet.merge import CellRange
from pandas import DataFrame


# ── small helpers ────────────────────────────────────────────────────────────
def _coerce(v: Any) -> Any:
    if v in (None, ""):
        return None
    if isinstance(v, (int, float)):
        return int(v) if isinstance(v, float) and v.is_integer() else v
    s = str(v).replace(",", "").strip()
    for cast in (int, float):
        try:
            return cast(s)
        except ValueError:
            continue
    return s


def _row_is_empty(ws: Worksheet, r: int, c0: int, c1: int) -> bool:
    return all(ws.cell(row=r, column=c).value in (None, "") for c in range(c0, c1 + 1))


def _merged_map(ws: Worksheet) -> Dict[tuple[int, int], CellRange]:
    """
    {(row, col): CellRange} for every cell *inside* a merged range
    (including the top-left anchor).
    """
    m: Dict[tuple[int, int], CellRange] = {}
    for rng in ws.merged_cells.ranges:
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                m[(r, c)] = rng
    return m


# ── main extractor ───────────────────────────────────────────────────────────
def extract_tables(
    file_path: str | Path,
    sheet_name: str,
    *,
    max_width: str | int = "AZ",
    debug: bool = False,
    sheet_skip_rows: Optional[list[int]] = None,
) -> Dict[str, List[Dict[str, Any]]]:
    """
    Row-by-row extractor that respects merged header cells.

    • Title row cell → check cell below → table starts.
    • Scan right for headers:
        – add header when cell has text
        – if cell is blank **but part of same merged range** as previous header,
          skip it (do not break).
        – stop when you hit a truly blank column (not merged with left header).
    • For each logical header we remember its *exact column index* so data keeps
      alignment even when headers span multiple columns.
    """
    max_col_limit = (
        column_index_from_string(max_width)
        if isinstance(max_width, str)
        else int(max_width)
    )

    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]
    if sheet_skip_rows:
        # delete these rows, biggest to smallest to avoid shifting
        for row in reversed(sheet_skip_rows):
            ws.delete_rows(row)
    max_row = ws.max_row
    merged_lookup = _merged_map(ws)

    tables: Dict[str, List[Dict[str, Any]]] = {}
    row = 1

    while row <= max_row:
        if debug:
            print(f"[ROW {row}] scanning")
        col = 1
        deepest_row_reached = row
        found_any_table = False

        while col <= max_col_limit:
            cell_val = ws.cell(row=row, column=col).value
            if debug:
                print(f"  → (R{row},C{col}) {repr(cell_val)}")
            if cell_val in (None, ""):
                col += 1
                continue

            # title candidate found – must have non-blank just below
            if ws.cell(row=row + 1, column=col).value in (None, ""):
                col += 1
                continue

            # ── TABLE DETECTED ────────────────────────────────────────────────
            found_any_table = True
            title = str(cell_val).strip()
            if debug:
                print(f"    >>> table '{title}' at R{row}C{col}")

            header_row = row + 1
            data_start_row = row + 2
            headers: List[str] = []
            header_cols: List[int] = []

            cur_col = col
            while cur_col <= max_col_limit:
                hv = ws.cell(row=header_row, column=cur_col).value
                if hv not in (None, ""):
                    headers.append(str(hv).strip())
                    header_cols.append(cur_col)
                    cur_col += 1
                    continue

                # hv is blank – is it part of a merge that started earlier?
                rng = merged_lookup.get((header_row, cur_col))
                if rng and rng.min_row == header_row and rng.min_col < cur_col:
                    # continuation of previous header – skip it
                    cur_col += 1
                    continue

                break  # truly blank – header block ends

            if not headers:  # should not happen
                col += 1
                continue

            last_header_col = cur_col - 1
            if debug:
                print(f"        headers: {headers} (cols {header_cols})")

            # ── DATA ROWS ─────────────────────────────────────────────────────
            data: List[Dict[str, Any]] = []
            r = data_start_row
            while r <= max_row and not _row_is_empty(ws, r, col, last_header_col):
                if debug:
                    print(f"          data row R{r}")
                rec = {
                    hdr: _coerce(ws.cell(row=r, column=header_cols[i]).value)
                    for i, hdr in enumerate(headers)
                }
                data.append(rec)
                r += 1

            # store, avoid duplicate keys
            unique_title = title
            n = 2
            while unique_title in tables:
                unique_title = f"{title} #{n}"
                n += 1
            tables[unique_title] = data
            if debug:
                print(f"    <<< finished '{unique_title}' with {len(data)} rows\n")

            deepest_row_reached = max(deepest_row_reached, r)
            col = last_header_col + 1  # resume scanning to the right

        # advance to next scan line
        row = deepest_row_reached if found_any_table else row + 1

    return tables


def convert_to_df_dict(tables: Dict[str, List[Dict[str, Any]]]) -> Dict[str, DataFrame]:

    data = {}

    for title, rows in tables.items():
        df = DataFrame(rows)
        data[title] = df

    return data
